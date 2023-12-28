#simple program for manipulating excel sheets 
from openpyxl import load_workbook

def process_workbook(filename): # filenames should be ".."
    wb = load_workbook(filename)
    sheet = wb["sheet_name"]
    
    for row in range(2, sheet.max_row + 1) : # 2 represents the second row of each column cause it is assumed the first row is used for titles... but if yours is without titles adjust as you like.
        cell1 = sheet.cell(row, 3) # digit represents specific column number so edit
        cell2 = sheet.cell(row, 1)
        
        a = cell1.value  # add as many cells and values as you like to manipulate
        b = cell2.value
       
        Total_price = int(a*b) # use int/float to allow these datasets bet manipulated mathematically... then use any mathematical symbol to automate (+ / - *)
        Total_price_cell = sheet.cell(row, 5) # choose row&col for the storing of each solution
        Total_price_cell.value = Total_price
# add as many formulas as you might need
        wb.save("...") #save file
       
