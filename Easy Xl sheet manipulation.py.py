#simple program for manipulating excel sheets 
from openpyxl import load_workbook

def process_workbook(filename): # filenames should be ".."
    wb = load_workbook(filename)
    sheet = wb["sheet_name"]
    
    for row in range(2, sheet.max_row + 1) :
        cell1 = sheet.cell(row, 3) # digit represents specific columber number so edit
        cell2 = sheet.cell(row, 1)
        
        a = cell1.value  # add as many cells and values as you like to manipulate
        b = cell2.value
       
        Total_price = int(a*b) # use int/float to allow these datasets bet manipulated mathematically 
        Total_price_cell = sheet.cell(row, 5) # choose row&col for the storing of each solution
        Total_price_cell.value = Total_price
        wb.save("I1.xlsx") #save file
       